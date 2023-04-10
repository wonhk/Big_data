from openpyxl import Workbook
import fun_movie
   
# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '순위'
write_ws['B1'] = '제목'
write_ws['C1'] = '평점'

m_list = fun_movie.get_dmv()
for m in m_list:
    #행 단위로 추가
    write_ws.append(m)

#셀 단위로 추가
#write_ws.cell(5, 5, '5행5열')

write_wb.save('함수로 만든 영화정보.xlsx')